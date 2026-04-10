"""
导出服务

处理表格数据的多格式导出（CSV、Excel、JSON）
"""
import io
import csv
import json
import logging
from datetime import datetime
from typing import List, Dict, Any, Tuple, Optional

# 顶部导入 openpyxl，提供清晰的依赖检测
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    Workbook = None

logger = logging.getLogger('table_parser.export_service')


class ExportError(Exception):
    """导出错误"""
    pass


class ExportService:
    """表格数据导出服务"""
    
    # Excel 样式定义（类级别，避免重复创建）
    _excel_styles_initialized = False
    _header_font = None
    _header_fill = None
    _header_alignment = None
    _thin_border = None
    
    @classmethod
    def _init_excel_styles(cls):
        """初始化 Excel 样式（延迟初始化）"""
        if cls._excel_styles_initialized or not EXCEL_AVAILABLE:
            return
        
        cls._header_font = Font(bold=True, color='FFFFFF')
        cls._header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cls._header_alignment = Alignment(horizontal='center', vertical='center')
        cls._thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cls._excel_styles_initialized = True
    
    @staticmethod
    def check_excel_support() -> bool:
        """检查是否支持 Excel 导出"""
        return EXCEL_AVAILABLE
    
    @staticmethod
    def export_csv(tables: List[Dict], table_index: int = 0) -> Tuple[str, str]:
        """
        导出单个表格为 CSV 格式
        
        Args:
            tables: 表格数据列表
            table_index: 要导出的表格索引
            
        Returns:
            (CSV 内容, 文件名)
            
        Raises:
            ExportError: 导出失败
        """
        if not tables:
            raise ExportError('没有可导出的表格数据')
        
        if table_index >= len(tables):
            raise ExportError(f'表格索引超出范围，共 {len(tables)} 个表格')
        
        table = tables[table_index]
        table_data = table.get('data', [])
        if not table_data:
            raise ExportError('表格数据为空')
        
        output = io.StringIO()
        writer = csv.writer(output)
        for row in table_data:
            writer.writerow(row)
        
        csv_content = output.getvalue()
        output.close()
        
        # 使用表格数据中的 global_index 生成文件名，确保一致性
        display_index = table.get('global_index', table_index) + 1
        filename = f'table_{display_index}.csv'
        logger.info(f"导出 CSV 成功: {filename}, {len(table_data)} 行")
        
        return csv_content, filename
    
    @classmethod
    def export_excel(cls, tables: List[Dict], table_index: int = 0, 
                     export_all: bool = False) -> Tuple[bytes, str]:
        """
        导出表格为 Excel 格式
        
        Args:
            tables: 表格数据列表
            table_index: 要导出的表格索引（export_all=False 时有效）
            export_all: 是否导出所有表格到一个文件
            
        Returns:
            (Excel 文件内容, 文件名)
            
        Raises:
            ExportError: 导出失败
        """
        if not EXCEL_AVAILABLE:
            raise ExportError('Excel 导出功能需要安装 openpyxl: pip install openpyxl')
        
        if not tables:
            raise ExportError('没有可导出的表格数据')
        
        # 初始化样式
        cls._init_excel_styles()
        
        wb = Workbook()
        
        if export_all:
            # 导出所有表格，每个表格一个 sheet
            wb.remove(wb.active)
            for idx, table in enumerate(tables):
                table_data = table.get('data', [])
                if not table_data:
                    continue
                
                # 使用 global_index 作为 sheet 名称，确保一致性
                display_index = table.get('global_index', idx) + 1
                ws = wb.create_sheet(title=f'表格{display_index}')
                cls._write_table_to_sheet(ws, table_data)
            
            filename = 'tables_all.xlsx'
        else:
            # 只导出指定表格
            if table_index >= len(tables):
                raise ExportError(f'表格索引超出范围，共 {len(tables)} 个表格')
            
            table = tables[table_index]
            table_data = table.get('data', [])
            if not table_data:
                raise ExportError('表格数据为空')
            
            # 使用 global_index 作为 sheet 名称和文件名，确保一致性
            display_index = table.get('global_index', table_index) + 1
            ws = wb.active
            ws.title = f'表格{display_index}'
            cls._write_table_to_sheet(ws, table_data)
            
            filename = f'table_{display_index}.xlsx'
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"导出 Excel 成功: {filename}")
        
        return output.getvalue(), filename
    
    @classmethod
    def _write_table_to_sheet(cls, ws, table_data: List[List[Any]]) -> None:
        """将表格数据写入 Excel 工作表"""
        for row_idx, row in enumerate(table_data, start=1):
            for col_idx, cell_value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value or '')
                cell.border = cls._thin_border
                cell.alignment = Alignment(vertical='center')
                
                # 首行样式（表头）
                if row_idx == 1:
                    cell.font = cls._header_font
                    cell.fill = cls._header_fill
                    cell.alignment = cls._header_alignment
        
        # 自动调整列宽
        for col_idx in range(1, len(table_data[0]) + 1 if table_data else 1):
            max_length = 0
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            
            for row_idx in range(1, len(table_data) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    max_length = max(max_length, cell_len)
                except:
                    pass
            
            ws.column_dimensions[column_letter].width = min(max_length + 4, 50)
    
    @staticmethod
    def export_json(tables: List[Dict], table_index: int = 0, 
                    export_all: bool = False) -> Tuple[str, str]:
        """
        导出表格为 JSON 格式
        
        Args:
            tables: 表格数据列表
            table_index: 要导出的表格索引（export_all=False 时有效）
            export_all: 是否导出所有表格
            
        Returns:
            (JSON 内容, 文件名)
            
        Raises:
            ExportError: 导出失败
        """
        if not tables:
            raise ExportError('没有可导出的表格数据')
        
        if export_all:
            export_data = {
                'total_tables': len(tables),
                'exported_at': datetime.now().isoformat(),
                'tables': []
            }
            for idx, table in enumerate(tables):
                # 使用 global_index 确保索引一致性
                display_index = table.get('global_index', idx) + 1
                export_data['tables'].append({
                    'index': display_index,
                    'global_index': table.get('global_index', idx),
                    'page': table.get('page'),
                    'page_table_index': table.get('page_table_index'),
                    'rows': len(table.get('data', [])),
                    'cols': len(table.get('data', [[]])[0]) if table.get('data') else 0,
                    'confidence': table.get('confidence', 0),
                    'data': table.get('data', [])
                })
            filename = 'tables_all.json'
        else:
            if table_index >= len(tables):
                raise ExportError(f'表格索引超出范围，共 {len(tables)} 个表格')
            
            table = tables[table_index]
            # 使用 global_index 确保索引一致性
            display_index = table.get('global_index', table_index) + 1
            export_data = {
                'table_index': display_index,
                'global_index': table.get('global_index', table_index),
                'page': table.get('page'),
                'page_table_index': table.get('page_table_index'),
                'exported_at': datetime.now().isoformat(),
                'rows': len(table.get('data', [])),
                'cols': len(table.get('data', [[]])[0]) if table.get('data') else 0,
                'confidence': table.get('confidence', 0),
                'data': table.get('data', [])
            }
            filename = f'table_{display_index}.json'
        
        json_content = json.dumps(export_data, ensure_ascii=False, indent=2)
        
        logger.info(f"导出 JSON 成功: {filename}")
        
        return json_content, filename
